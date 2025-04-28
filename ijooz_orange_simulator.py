import pandas as pd
import datetime
from io import BytesIO
import zipfile
import os
import tempfile
import streamlit as st
from openpyxl import load_workbook
from openpyxl.chart import LineChart, BarChart, Reference

# === 仓库容量配置 ===
warehouse_capacities = {
    'Singapore': 16,
    'Tokyo': 16,
    'Osaka': 4.5,
    'Nagoya': 5,
    'Fukuoka': 5
}

# === 图表函数 ===
def add_charts_to_workbook(wb):
    if "Daily Inventory" not in wb.sheetnames or "Container Schedule" not in wb.sheetnames:
        return

    sched_df = pd.DataFrame(wb["Container Schedule"].values)
    headers = sched_df.iloc[0].tolist()
    sched_df = sched_df[1:]
    sched_df.columns = headers
    if "开始使用时间" in sched_df.columns:
        sched_df = sched_df.sort_values(by="开始使用时间")

    wb.remove(wb["Container Schedule"])
    ws_new = wb.create_sheet("Container Schedule")
    for r_idx, row in enumerate([headers] + sched_df.values.tolist(), 1):
        for c_idx, val in enumerate(row, 1):
            ws_new.cell(row=r_idx, column=c_idx, value=val)

    inv_ws = wb["Daily Inventory"]
    inv_headers = [cell.value for cell in inv_ws[1]]
    inv_data = []
    for row in inv_ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(inv_headers, row))
        for col in ["IJOOZ 仓库库存（单位）", "总库存（单位）"]:
            if col in row_dict and isinstance(row_dict[col], (int, float)):
                row_dict[col] = round(row_dict[col], 1)
        inv_data.append(row_dict)

    wb.remove(inv_ws)
    new_inv_ws = wb.create_sheet("Daily Inventory")
    for c_idx, col in enumerate(inv_headers, 1):
        new_inv_ws.cell(row=1, column=c_idx, value=col)
    for r_idx, row in enumerate(inv_data, 2):
        for c_idx, col in enumerate(inv_headers, 1):
            new_inv_ws.cell(row=r_idx, column=c_idx, value=row.get(col))

    chart_sheet = wb.create_sheet("Charts")

    chart1 = LineChart()
    chart1.title = "每日库存趋势"
    chart1.height = 10
    chart1.width = 20
    chart1.x_axis.title = "日期"
    chart1.x_axis.numFmt = "yyyy-mm-dd"
    chart1.y_axis.title = "库存单位数"
    data = Reference(new_inv_ws, min_col=2, max_col=3, min_row=1, max_row=new_inv_ws.max_row)
    categories = Reference(new_inv_ws, min_col=1, min_row=2, max_row=new_inv_ws.max_row)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(categories)
    chart_sheet.add_chart(chart1, "A1")

    sched_ws = wb["Container Schedule"]
    chart2 = BarChart()
    chart2.title = "每柜生命周期（天）"
    chart2.height = 10
    chart2.width = 20
    po_col, life_col = 1, 1
    for col in range(1, sched_ws.max_column + 1):
        if sched_ws.cell(1, col).value == "PO":
            po_col = col
        if sched_ws.cell(1, col).value == "生命周期（天）":
            life_col = col
    data2 = Reference(sched_ws, min_col=life_col, min_row=1, max_row=sched_ws.max_row)
    categories2 = Reference(sched_ws, min_col=po_col, min_row=2, max_row=sched_ws.max_row)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(categories2)
    chart_sheet.add_chart(chart2, "A20")

# === 单仓库模拟函数 ===
def run_simulation(file, warehouse_name):
    if warehouse_name not in warehouse_capacities:
        raise ValueError(f"未定义仓库容量：{warehouse_name}")
    ijooz_capacity = warehouse_capacities[warehouse_name]

    xls = pd.ExcelFile(file)
    container_sheet = f"Container-{warehouse_name}"
    usage_sheet = f"weekly usage-{warehouse_name}"
    if container_sheet not in xls.sheet_names or usage_sheet not in xls.sheet_names:
        raise ValueError(f"Excel中未找到sheet：{container_sheet} 或 {usage_sheet}")

    container_df = xls.parse(container_sheet)
    weekly_usage_df = xls.parse(usage_sheet)
    weekly_usage_df[['year', 'week_number']] = weekly_usage_df['week'].str.extract(r'(\d{4})WK(\d{2})').astype(int)
    weekly_usage_df['monday'] = pd.to_datetime(
        weekly_usage_df['year'].astype(str) + weekly_usage_df['week_number'].astype(str) + '1', format='%G%V%u'
    )

    daily_usage_records = []
    for _, row in weekly_usage_df.iterrows():
        daily_usage = row['用量'] / 7
        for i in range(7):
            daily_usage_records.append({
                'date': row['monday'] + pd.Timedelta(days=i),
                'daily_usage': daily_usage
            })
    daily_usage_df = pd.DataFrame(daily_usage_records)

    container_df['HARVEST DAY'] = pd.to_datetime(container_df['HARVEST DAY'])
    container_df = container_df.sort_values(by='HARVEST DAY').reset_index(drop=True)

    today = pd.Timestamp(datetime.date.today())
    containers = []
    for idx, row in container_df.iterrows():
        eta = pd.to_datetime(row['ETA DATE']) if pd.notnull(row['ETA DATE']) else None
        containers.append({
            'index': idx,
            'PO': row['PO'],
            'Vessel': row.get('Vessel', ''),
            'harvest_day': pd.to_datetime(row['HARVEST DAY']),
            'eta': eta,
            'unit': float(row['单位']),
            'in_ext_date': None,
            'in_ijooz_date': today if eta is None else None,
            'start_use': None,
            'end_use': None,
            'used': 0,
            'can_enter_date': eta + pd.Timedelta(days=3) if eta is not None else today
        })

    ijooz_storage = [c for c in containers if c['in_ijooz_date'] is not None]
    external_storage = []
    used_capacity = sum(c['unit'] for c in ijooz_storage)
    inventory_log = []
    date_range = pd.date_range(start=today, end=daily_usage_df['date'].max())

    for day in date_range:
        used_today = []
        for c in containers:
            if c['in_ijooz_date'] is None and c['in_ext_date'] is None:
                if c['can_enter_date'] <= day:
                    if used_capacity + c['unit'] <= ijooz_capacity:
                        c['in_ijooz_date'] = day
                        ijooz_storage.append(c)
                        used_capacity += c['unit']
                    else:
                        c['in_ext_date'] = day
                        external_storage.append(c)

        external_storage.sort(key=lambda x: x['in_ext_date'])
        for c in external_storage[:]:
            if used_capacity + c['unit'] <= ijooz_capacity:
                c['in_ijooz_date'] = day
                ijooz_storage.append(c)
                used_capacity += c['unit']
                external_storage.remove(c)

        day_usage = daily_usage_df.loc[daily_usage_df['date'] == day, 'daily_usage'].sum()
        day_usage_original = day_usage
        while day_usage > 0 and ijooz_storage:
            c = ijooz_storage[0]
            if c['in_ijooz_date'] is not None and day < c['in_ijooz_date']:
                break
            remaining = c['unit'] - c['used']
            if c['start_use'] is None:
                c['start_use'] = day
            use_now = min(day_usage, remaining)
            c['used'] += use_now
            day_usage -= use_now
            if c['used'] == c['unit']:
                c['end_use'] = day
                used_capacity -= c['unit']
                ijooz_storage.pop(0)
            used_today.append(c['PO'])

        inventory_log.append({
            '日期': day,
            'IJOOZ 仓库库存（单位）': round(sum(c['unit'] - c['used'] for c in ijooz_storage), 1),
            '外部冷库库存（整柜数）': len(external_storage),
            '当天使用的货柜 PO': ', '.join(set(used_today)),
            '使用柜数量': len(set(used_today)),
            '总库存（单位）': round(sum(c['unit'] - c['used'] for c in ijooz_storage) + sum(c['unit'] for c in external_storage), 1),
            '运输中（单位）': round(sum(c['unit'] for c in containers if c['eta'] and c['eta'].date() > day.date()), 1),
            'daily_usage': round(day_usage_original, 2)
        })

    schedule_df = pd.DataFrame([{
        'Vessel': c['Vessel'],
        'PO': c['PO'],
        'Harvest Day': c['harvest_day'],
        'ETA': c['eta'],
        '单位': c['unit'],
        '进外面冷库时间': c['in_ext_date'],
        '进IJOOZ仓库时间': c['in_ijooz_date'],
        '开始使用时间': c['start_use'],
        '使用完的时间': c['end_use'],
        '生命周期（天）': (c['start_use'] - c['harvest_day']).days if c['start_use'] else None
    } for c in containers])

    inventory_df = pd.DataFrame(inventory_log)
    inventory_df['日期'] = pd.to_datetime(inventory_df['日期']).dt.strftime('%Y-%m-%d')
    schedule_df['Harvest Day'] = pd.to_datetime(schedule_df['Harvest Day']).dt.strftime('%Y-%m-%d')

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        schedule_df.to_excel(writer, index=False, sheet_name="Container Schedule")
        inventory_df.to_excel(writer, index=False, sheet_name="Daily Inventory")

    output.seek(0)
    wb = load_workbook(output)

    # 添加 Charts
    add_charts_to_workbook(wb)

    # 添加单仓库 Weekly Summary
    inventory_df["日期"] = pd.to_datetime(inventory_df["日期"])
    inventory_df["year"] = inventory_df["日期"].dt.isocalendar().year
    inventory_df["week"] = inventory_df["日期"].dt.isocalendar().week
    weekly_summary = inventory_df.groupby(["year", "week"]).agg({
        "IJOOZ 仓库库存（单位）": "mean",
        "外部冷库库存（整柜数）": "mean",
        "使用柜数量": "sum",
        "运输中（单位）": "mean",
        "daily_usage": "sum"
    }).reset_index()

    with pd.ExcelWriter(output, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        weekly_summary.to_excel(writer, index=False, sheet_name=f"{warehouse_name} Weekly Summary")

    wb.save(output)
    output.seek(0)

    return inventory_df, output


# === 批量生成 + 打包 zip ===
def run_all_simulations(file):
    st.session_state["run_all_mode"] = True

    xls = pd.ExcelFile(file)
    available_warehouses = [name.replace("Container-", "") for name in xls.sheet_names if name.startswith("Container-")]

    all_inventory_dfs = []
    regional_weeklies = {}

    with tempfile.TemporaryDirectory() as tmpdirname:
        excel_paths = []

        for wh in available_warehouses:
            try:
                inventory_df, sim_output = run_simulation(file, wh)
                all_inventory_dfs.append(inventory_df)
                regional_weeklies[wh] = inventory_df.copy()

                filename = f"{wh}_simulation.xlsx"
                file_path = os.path.join(tmpdirname, filename)
                with open(file_path, "wb") as f:
                    f.write(sim_output.read())
                excel_paths.append(file_path)

            except Exception as e:
                st.warning(f"⚠️ 仓库 {wh} 模拟失败：{e}")

        # === 汇总日本 Daily Inventory
        if all_inventory_dfs:
            combined = pd.concat(all_inventory_dfs)
            combined["日期"] = pd.to_datetime(combined["日期"])

            japan_grouped = combined.groupby("日期", as_index=False).agg({
                "IJOOZ 仓库库存（单位）": "sum",
                "外部冷库库存（整柜数）": "sum",
                "使用柜数量": "sum",
                "运输中（单位）": "sum",
                "daily_usage": "sum",
                "当天使用的货柜 PO": lambda x: ', '.join(filter(None, map(str, x)))
            })

            japan_path = os.path.join(tmpdirname, "Japan_Daily_Inventory.xlsx")
            with pd.ExcelWriter(japan_path, engine="openpyxl") as writer:
                japan_grouped.to_excel(writer, index=False, sheet_name="Japan Daily Inventory")

                # === 日本 Weekly Summary
                japan_grouped["year"] = japan_grouped["日期"].dt.isocalendar().year
                japan_grouped["week"] = japan_grouped["日期"].dt.isocalendar().week
                japan_weekly = japan_grouped.groupby(["year", "week"]).agg({
                    "IJOOZ 仓库库存（单位）": "mean",
                    "外部冷库库存（整柜数）": "mean",
                    "使用柜数量": "sum",
                    "运输中（单位）": "mean",
                    "daily_usage": "sum"
                }).reset_index()
                japan_weekly.to_excel(writer, index=False, sheet_name="Japan Weekly Summary")

                # === 每个城市的 Weekly Summary
                for wh, region_df in regional_weeklies.items():
                    region_df["日期"] = pd.to_datetime(region_df["日期"])
                    region_df["year"] = region_df["日期"].dt.isocalendar().year
                    region_df["week"] = region_df["日期"].dt.isocalendar().week
                    region_weekly = region_df.groupby(["year", "week"]).agg({
                        "IJOOZ 仓库库存（单位）": "mean",
                        "外部冷库库存（整柜数）": "mean",
                        "使用柜数量": "sum",
                        "运输中（单位）": "mean",
                        "daily_usage": "sum"
                    }).reset_index()
                    region_weekly.to_excel(writer, index=False, sheet_name=f"{wh} Weekly Summary")

            excel_paths.append(japan_path)

        # === 打包所有文件到 zip
        zip_output = BytesIO()
        with zipfile.ZipFile(zip_output, "w") as zipf:
            for path in excel_paths:
                arcname = os.path.basename(path)
                zipf.write(path, arcname=arcname)

        zip_output.seek(0)
        return zip_output

# === 页面设置 ===
st.set_page_config(page_title="IJOOZ 仓库模拟器", page_icon="🍊", layout="centered")
st.markdown('<p class="title-text">🍊 IJOOZ 仓库模拟器</p>', unsafe_allow_html=True)
st.markdown('<p class="subtitle-text">上传仓库使用计划 Excel 文件，自动计算库存及生命周期，并生成图表。</p>', unsafe_allow_html=True)
st.markdown("---")

warehouse_options = list(warehouse_capacities.keys())
warehouse_options.insert(0, '全部仓库')
warehouse_name = st.selectbox("📍 选择仓库地点", warehouse_options, index=0)
uploaded_file = st.file_uploader("📤 上传 Excel 文件", type=["xlsx", "xls"])

# === 主入口 ===
if uploaded_file and st.button("🚀 运行模拟"):
    try:
        today_str = datetime.date.today().strftime('%Y-%m-%d')
        with st.spinner("模拟进行中，请稍候..."):
            if warehouse_name == '全部仓库':
                output_zip = run_all_simulations(uploaded_file)
                filename = f"IJOOZ_Simulation_ALL_{today_str}.zip"
                st.success("✅ 所有仓库模拟完成！点击下方按钮下载所有结果：")
                st.download_button("📦 下载 ZIP 文件", data=output_zip, file_name=filename, mime="application/zip")
            else:
                _, output_excel = run_simulation(uploaded_file, warehouse_name)
                filename = f"IJOOZ_Simulation_{warehouse_name}_{today_str}.xlsx"
                st.success("✅ 仓库模拟完成！点击下方按钮下载结果：")
                st.download_button("📥 下载 Excel 文件", data=output_excel, file_name=filename)
    except Exception as e:
        st.error(f"❌ 出错了：{str(e)}")
